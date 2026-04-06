"""Generación de recibo Excel (.xlsx) a partir de datos y productos."""
from __future__ import annotations

import re
from datetime import date
from pathlib import Path

from openpyxl import load_workbook

from generator.product import Product


def fill_recibo(
    data: dict,
    products: list[Product],
    template_path: str = "pruebas/RECIBO.xlsx",
    output_dir: str = "pruebas",
) -> Path:
    """Genera un recibo Excel insertando los productos en la plantilla RECIBO.xlsx."""
    wb = load_workbook(template_path)
    ws = wb.active

    if ws is None:
        raise ValueError("No se pudo cargar la hoja de cálculo")

    combine = [(2, 3), (4, 5), (6, 7)]
    valid_col = [1, 2, 3, 5, 7, 9, 10]

    for i, product in enumerate(products):
        index = 3 + i
        ws.insert_rows(index)

        for start, end in combine:
            ws.merge_cells(
                start_row=index, start_column=start,
                end_row=index, end_column=end,
            )

        for col, value in enumerate(product.get_list()):
            ws.cell(row=index, column=valid_col[col]).value = value

    direccion_slug = re.sub(r'[^\w\s-]', '', data.get("direccion", "documento")).strip().replace(" ", "_")
    timestamp = date.today().strftime("%Y%m%d")
    output_path = Path(output_dir) / f"RECIBO_{direccion_slug}_{timestamp}.xlsx"
    wb.save(output_path)
    return output_path
